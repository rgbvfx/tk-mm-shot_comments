# Copyright (c) 2013 Shotgun Software Inc.
# Mind Machine customized

import os
import re
import sys
import sgtk
from openpyxl import load_workbook
from sgtk.platform.qt import QtCore, QtGui

# standard toolkit logger
logger = sgtk.platform.get_logger(__name__)


def show_dialog(app_instance):
    """Show the main dialog window."""
    app_instance.engine.show_dialog("Shot Comments App", app_instance, AppDialog)


class AppDialog(QtGui.QWidget):
    """Main application dialog window"""
    
    def __init__(self):
        """
        Constructor
        """
        # first, call the base class and let it do its thing.
        QtGui.QWidget.__init__(self)
        
        logger.info('Launching Shot Comments app...')

        # create widgets
        self.button_file_open = QtGui.QPushButton('Select Excel Files')
        self.button_shotgun_import = QtGui.QPushButton('Shotgun Import')
        self.label_status = QtGui.QLabel()
        self.logo_example = QtGui.QLabel()
        self.progress_bar = QtGui.QProgressBar()
        self.table = QtGui.QTableWidget()

        # layout
        self.vertical_layout_1 = QtGui.QVBoxLayout()
        self.vertical_layout_2 = QtGui.QVBoxLayout()
        self.horizontal_layout_1 = QtGui.QHBoxLayout()
        self.vertical_layout_1.addWidget(self.logo_example)
        self.vertical_layout_1.addWidget(self.button_file_open)
        self.vertical_layout_1.addWidget(self.button_shotgun_import)
        spacerItem = QtGui.QSpacerItem(20, 40, QtGui.QSizePolicy.Minimum, QtGui.QSizePolicy.Expanding)
        self.vertical_layout_1.addItem(spacerItem)
        self.vertical_layout_2.addWidget(self.progress_bar)
        self.vertical_layout_2.addWidget(self.table)
        self.vertical_layout_2.addWidget(self.label_status)
        self.horizontal_layout_1.addLayout(self.vertical_layout_1)
        self.horizontal_layout_1.addLayout(self.vertical_layout_2)
        # set dialog layout
        self.setLayout(self.horizontal_layout_1)

        # size
        self.button_file_open.setMinimumWidth(80)
        self.button_file_open.setMaximumWidth(100)
        size_policy = QtGui.QSizePolicy(QtGui.QSizePolicy.Fixed, QtGui.QSizePolicy.Fixed)
        size_policy.setHorizontalStretch(0)
        size_policy.setVerticalStretch(0)
        size_policy.setHeightForWidth(self.logo_example.sizePolicy().hasHeightForWidth())
        self.logo_example.setSizePolicy(size_policy)
        self.logo_example.setMaximumSize(QtCore.QSize(256, 256))
        self.logo_example.setText("")
        self.logo_example.setPixmap(QtGui.QPixmap(":/res/sg_logo.png"))
        self.logo_example.setAlignment(QtCore.Qt.AlignCenter)
        self.logo_example.setObjectName("logo_example")

        # connect buttons
        self.button_file_open.clicked.connect(self._select_files)
        self.button_shotgun_import.clicked.connect(self._shotgun_import)

        # self._app.context
        # self._app.context.user
        # self._app.engine
        # self._app.project
        # self._app.shotgun   etc...
        self._app = sgtk.platform.current_bundle()

        # data
        self.default_color = None
        self.header_list = None
        self.last_edl_file_path = None
        self.project = self._app.context.project
        self.project_name = self.project['name']
        self.received_error = False
        self.table_rows_need_to_be_deleted = False
        self.user = self._app.context.user
        self.user_first_name = 'User'
        if self.user:
            self.user_first_name = self.user['name'].split()[0]
        self.wb = None
        self.ws = None

        # set-up gui
        self.button_shotgun_import.hide()
        self.progress_bar.hide()
        self.progress_bar.setValue(0)
        msg = 'Hi {}!   Please select one or more excel files...'.format(self.user['name'])
        self.label_status.setText(unicode(msg))

        # sg connection
        self.sg = self._app.shotgun

        # thread placeholder
        self._thread = None

        msg = 'Shot Comments import app initialize by {}'.format(self.user['name'])
        logger.info(msg)

    def _add_data_to_table(self, data):
        # add header row and data rows

        # create header row, if this is the first we add data
        if self.table.rowCount() == 0:
            self.table.setColumnCount(len(self.header_list))
            for col, header_name in enumerate(self.header_list):
                item = QtGui.QTableWidgetItem()
                item.setText(header_name)
                item.setFlags(QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable)  # not ItemIsEditable
                self.table.setHorizontalHeaderItem(col, item)
                self.table.show()

        # add data
        for data_row_dict in data:
            row = self.table.rowCount()
            self.table.setRowCount(row + 1)
            for col, header_name in enumerate(self.header_list):
                item = QtGui.QTableWidgetItem()
                item.setText(data_row_dict[header_name])
                item.setFlags(QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable)  # not ItemIsEditable
                self.table.setItem(row, col, item)

        # get default color
        if self.table.rowCount() > 0 and self.table.columnCount() > 0:
            item = self.table.item(0, 0)
            if item:
                self.default_color = item.background()

        msg = 'I like the way you did that, {}!'.format(self.user_first_name)
        self.label_status.setText(unicode(msg))
        self.table.update()

    def _delete_table_rows(self):
        """Recursively delete table rows"""
        for row in range(self.table.rowCount()):
            self.table.removeRow(row)
        if self.table.rowCount():
            self._delete_table_rows()
        else:
            # they are now deleted
            self.table_rows_need_to_be_deleted = False

    @staticmethod
    def _get_headers():
        return ['VFX Shot', 'Internal Comment Date', 'Comment Version',	'Shot Status', 'VFX Comments', 'Vendor']

    def _open_file(self, file_path):
        self.wb = load_workbook(filename=file_path)
        self.ws = self.wb.active
        self._parse_excel_file()

    def _parse_excel_file(self):
        header_dict = dict()
        header_list = list()
        header_row = None
        found_comment = False
        found_shot_code = False
        found_comment_version = False
        found_vendor = False

        # get header row data
        for row in self.ws.rows:
            for cell in row:
                cell_col = int(cell.column)
                header_dict[cell_col] = {'val': cell.value, 'row': int(cell.row), 'col_index': int(cell.column)}
                header_list.append(cell.value)
                if cell.value == 'VFX Shot':
                    found_shot_code = True
                    header_row = cell.row
                elif cell.value == 'Comment Version':
                    found_comment_version = True
                elif cell.value == 'VFX Comments':
                    found_comment = True
                elif cell.value == 'Vendor':
                    found_vendor = True
            if found_shot_code and found_comment and found_comment_version and found_vendor:
                break

        if not header_row:
            msg = 'ERROR: could not parse excel file - failed to get header row. Please close application.'
            logger.info(msg)
            self.label_status.setText(msg)
            raise RuntimeError

        row_data_start = header_row + 1
        row_data_end = self.ws.max_row + 1

        col_keys = header_dict.keys()
        col_keys.sort()
        col_start = col_keys[0]
        col_end = col_keys[-1] + 1

        data = list()

        # put each row into data dict
        for row in range(row_data_start, row_data_end):
            data_row_dict = dict()
            for col in range(col_start, col_end):
                cell = self.ws.cell(row, col)
                header_name = header_dict[col]['val']
                data_row_dict[header_name] = unicode(cell.value)
            data.append(data_row_dict)

        self.header_list = header_list
        self._add_data_to_table(data)

    def _select_files(self):

        # if we are starting over, delete table rows from last run
        if self.table_rows_need_to_be_deleted:
            self._delete_table_rows()

        start_path = '~'

        if self.last_edl_file_path:
            start_path = self.last_edl_file_path
        else:
            # 'darwin'  'linux'  'win32'
            if sys.platform == 'win32':
                start_path = 'C:\\'

        try:
            result = QtGui.QFileDialog().getOpenFileNames(self, u'Select Files', start_path, "*.xlsx")
        except IOError:
            msg = 'ERROR: failed to get items from file dialog.'
            logger.info(msg)
            self.label_status.setText(msg)
            return

        if result and result[0]:
            # check file path
            xl_file_list = result[0]
            for xl_file_path in xl_file_list:
                if not os.path.exists(xl_file_path):
                    msg = 'ERROR: excel file path does not exist: {}'.format(xl_file_path)
                    logger.info(msg)
                    self.label_status.setText(msg)
                    return
                # success
                self.last_edl_file_path = os.path.dirname(xl_file_list[0])
                logger.info('Selected file: {}'.format(xl_file_path))
                self._open_file(xl_file_path)
        else:
            msg = 'WARNING: no path from file dialog. User may have canceled.'
            logger.info(msg)
            return

        self.button_file_open.hide()
        self.button_shotgun_import.show()

    def _set_row_color(self, row, color_name):

        if color_name == 'bright green':
            row_color = QtGui.QColor(240, 255, 220)
        elif color_name == 'green':
            row_color = QtGui.QColor(200, 255, 200)
        elif color_name == 'blue':
            row_color = QtGui.QColor(140, 150, 220)
        elif color_name == 'light blue':
            row_color = QtGui.QColor(240, 240, 255)
        elif color_name == 'red':
            row_color = QtGui.QColor(255, 96, 96)
        elif color_name == 'dark red':
            row_color = QtGui.QColor(180, 16, 16)
        elif color_name == 'violet':
            row_color = QtGui.QColor(240, 230, 255)
        elif color_name == 'ultra_violet':
            row_color = QtGui.QColor(230, 210, 255)
        elif color_name == 'yellow':
            row_color = QtGui.QColor(128, 128, 48)
        elif color_name == 'gray':
            row_color = QtGui.QColor(96, 96, 96)
        elif color_name == 'default':
            if self.default_color:
                row_color = self.default_color
            else:
                row_color = QtGui.QColor(96, 96, 96)
        else:  # no match
            row_color = QtGui.QColor(255, 255, 255)

        for col in range(self.table.columnCount()):
            item = self.table.item(row, col)
            if item:
                item.setBackground(row_color)

    def _shotgun_import(self):
        """Create data and send to thread for processing.
        :return: None
        """
        self.button_shotgun_import.hide()

        shot_data_list = list()

        for row in range(self.table.rowCount()):

            # pull data from table
            data_row_dict = dict()
            for col, header_name in enumerate(self.header_list):
                item = self.table.item(row, col)
                data_row_dict[header_name] = item.text()

            # reformat into *shot_data_dict* for creation of note
            shot_data_dict = dict()
            shot_status = data_row_dict['Shot Status']
            shot_version = data_row_dict['Comment Version']
            comment_date = data_row_dict['Internal Comment Date']
            comment_text = '{}\n\n'.format(data_row_dict['VFX Comments'])
            comment_text += 'Comment Version: {}\n'.format(shot_version)
            comment_text += 'Internal Comment Date: {}\n'.format(comment_date)
            comment_text += 'Shot Status: {}\n'.format(shot_status)
            comment_text += 'Vendor: {}'.format(data_row_dict['Vendor'])
            subject_text = '{} || {}'.format(shot_version, shot_status)
            shot_data_dict['shot_name'] = data_row_dict['VFX Shot']
            shot_data_dict['shot_version'] = shot_version
            shot_data_dict['comment_text'] = comment_text
            shot_data_dict['subject_text'] = subject_text
            shot_data_dict['row'] = row

            # append to list
            shot_data_list.append(shot_data_dict)

        self.progress_bar.setMaximum(len(shot_data_list))
        self.progress_bar.update()
        self.progress_bar.show()
        self.button_shotgun_import.hide()

        # close gui connection to shotgun, we'll use thread connection
        self.sg.close()
        self.sg = None

        msg = 'Starting thread'
        logger.info(msg)

        # send *shot_data_list* to thread for processing
        self._thread = SGProcessThread(shot_data_list=shot_data_list)
        self._thread.finished.connect(self._thread_notify_finish)
        self._thread.signal_from_thread.connect(self._thread_receive)
        self._thread.start()

    def _start_over(self):
        msg = 'Shot comments import complete'
        logger.info(msg)
        msg = 'Wow, {}! You really know what you\'re doing. Nice job!'.format(self.user_first_name)
        if self.received_error:
            msg = 'There was an error. Sorry about that, {}. Please report the problem to department of errors... '.format(self.user_first_name)
        self.label_status.setText(msg)
        self.progress_bar.hide()
        self.button_shotgun_import.hide()
        self.button_file_open.show()
        self.received_error = False
        self.sg = self._app.shotgun
        self.table_rows_need_to_be_deleted = True
        self.update()

    def _thread_receive(self, shot_code, msg, row):
        # receive message from thread
        self.progress_bar.setValue(row + 1)
        self.progress_bar.update()
        if msg == 'imported':
            self._set_row_color(row, 'green')
        elif msg == 'error':
            self._set_row_color(row, 'dark red')
            self.received_error = True
        elif msg == 'test':
            self._set_row_color(row, 'yellow')
        elif msg == 'note_data':
            self._set_row_color(row, 'light blue')
        elif msg == 'no_shot':
            self._set_row_color(row, 'red')
        # set next row to bright green
        if row < self.table.rowCount():
            self._set_row_color(row + 1, 'bright green')

        message = 'thread processed row {}:  {}  {}'.format(row, shot_code, msg)
        logger.info(message)
        self.update()

    def _thread_notify_finish(self):
        self._thread = None
        logger.info('Thread finished')
        self._start_over()

    def _thread_send(self, shot_data_list):
        # send shot list to thread
        self.app_signals.from_gui.emit(shot_data_list)


class SGProcessThread(QtCore.QThread):
    """Thread to create/import comments in Shotgun."""

    # note signal must be created before thread initialization
    signal_from_thread = QtCore.Signal(str, str, int)

    def __init__(self, shot_data_list):
        """Initialize thread.
        :param shot_data_list: list of dictionaries
        """
        QtCore.QThread.__init__(self)
        self.shot_data_list = shot_data_list
        self._app = sgtk.platform.current_bundle()
        self.project = self._app.context.project
        self.rgx_version = re.compile('^[A-Z]*[0-9]*_(v[0-9]*)_*')
        self.sg = self._app.shotgun
        self.user = self._app.context.user

        # TODO: CHECK TEST MODE BEFORE RELEASE
        self.test = False

    def __del__(self):
        self.wait()

    def run(self):
        """Loop through shot_data and process each note.
        :return: None
        """
        for shot_data in self.shot_data_list:
            self.process_shot_data(shot_data)

        self.sg.close()
        self.sg = None

    def process_shot_data(self, shot_data):
        """Create note in shotgun.
        :return: None
        """

        shot_code = shot_data['shot_name']
        row_number = shot_data['row']

        # create note data
        note_data = self._create_note_data(shot_data)

        if not note_data:
            self.signal_from_thread.emit(shot_code, 'no_shot', row_number)
            return

        if self.test:
            return

        # create note
        try:
            result = self.sg.create('Note', note_data)
            # success
            if result:
                status = 'imported'
                self.signal_from_thread.emit(shot_code, status, row_number)
            else:
                status = 'error'
                self.signal_from_thread.emit(shot_code, status, row_number)
                return
        except:
            status = 'error'
            self.signal_from_thread.emit(shot_code, status, row_number)
            return

    def _create_note_data(self, shot_data):
        """Create note data.
        :return: dict
        """
        shot_code = shot_data['shot_name']
        version_code = shot_data['shot_version']
        comment_text = shot_data['comment_text']
        subject_text = shot_data['subject_text']

        # find shot in shotgun
        filters = [['code', 'is', shot_code], ['project', 'is', self.project]]
        shot = self.sg.find_one('Shot', filters, ['code', 'sg_cut_duration', 'sg_versions'])

        # no shot?
        if not shot:
            return None

        # link to shot
        note_links = [{'type': 'Shot', 'id': shot['id']}]

        found_version = None

        # find the version the note will link to
        for version in shot['sg_versions']:
            if version_code in version['name']:
                found_version = version
                break

        # if that didn't work try matching just the version number eg:  v0007
        if not found_version:
            match_list = self.rgx_version.findall(version_code)
            if match_list:
                version_string = match_list[0]
                for version in shot['sg_versions']:
                    if version_string in version['name']:
                        found_version = version
                        break

        # link to version
        if found_version:
            note_links.append(found_version)

        # create the note data
        note_data = dict(project=self.project,
                         content=comment_text,
                         note_links=note_links,
                         subject=subject_text,
                         sg_note_type='Client')

        return note_data
